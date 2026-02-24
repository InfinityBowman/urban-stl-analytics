#!/usr/bin/env python3
"""
clean_data.py — Process raw datasets into frontend-ready JSON/GeoJSON.

Reads from python/data/raw/ (populated by fetch_raw.py) and writes
processed files to public/data/.

Usage:
  cd python/
  uv run python scripts/fetch_raw.py   # download first (if not already done)
  uv run python scripts/clean_data.py  # then process

Set DATA_YEAR env var to change the target year (default: 2025).
"""

import csv
import io
import json
import math
import os
import re
import shutil
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    requests = None  # only needed for weather API

try:
    import shapefile  # pyshp
except ImportError:
    sys.exit("Missing dependency: uv sync")

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: uv sync")

# ── Config ───────────────────────────────────────────────────────────────────

PYTHON_DIR = Path(__file__).resolve().parent.parent  # python/
ROOT = PYTHON_DIR.parent  # repo root
RAW_DIR = PYTHON_DIR / "data" / "raw"
OUT_DIR = ROOT / "public" / "data"

# Load .env from repo root
_dotenv = ROOT / ".env"
if _dotenv.exists():
    for line in _dotenv.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip())

YEAR = int(os.environ.get("DATA_YEAR", "2025"))
ACS_YEAR = int(os.environ.get("ACS_YEAR", "2022"))  # ACS data lags ~2 years

STL_COUNTY_FIPS = "29510"

# ── Helpers ──────────────────────────────────────────────────────────────────

def log(msg: str):
    print(f"  → {msg}")


def require_raw(path: Path, name: str):
    """Exit with a helpful message if raw data is missing."""
    if not path.exists():
        sys.exit(f"Missing raw data: {path}\nRun `uv run python scripts/fetch_raw.py` first.")


def web_mercator_to_lnglat(x: float, y: float) -> tuple[float, float]:
    """Convert Web Mercator (EPSG:3857) coordinates to lon/lat (EPSG:4326)."""
    lng = x * 180.0 / 20037508.34
    lat = math.atan(math.exp(y * math.pi / 20037508.34)) * 360.0 / math.pi - 90.0
    return lng, lat


def shapefile_to_geojson(shp_path: str) -> dict:
    """Convert a shapefile to GeoJSON FeatureCollection."""
    sf = shapefile.Reader(shp_path)
    fields = [f[0] for f in sf.fields[1:]]
    features = []
    for sr in sf.shapeRecords():
        props = dict(zip(fields, sr.record))
        for k, v in props.items():
            if isinstance(v, (bytes, bytearray)):
                props[k] = v.decode("utf-8", errors="replace")
        geom = sr.shape.__geo_interface__
        features.append({"type": "Feature", "properties": props, "geometry": geom})
    return {"type": "FeatureCollection", "features": features}


def safe_int(v, default=0):
    try:
        return int(v) if v is not None and str(v).strip().upper() != "NULL" else default
    except (ValueError, TypeError):
        return default


def safe_float(v, default=0.0):
    try:
        return float(v) if v is not None and str(v).strip().upper() != "NULL" else default
    except (ValueError, TypeError):
        return default


# ── 1. CSB 311 Data ──────────────────────────────────────────────────────────

def process_csb() -> None:
    """Process CSB 311 complaint CSVs from raw data."""
    csb_dir = RAW_DIR / "csb"
    require_raw(csb_dir, "CSB")

    csv_files = list(csb_dir.rglob("*.csv"))
    if not csv_files:
        sys.exit("No CSV files found in raw/csb/")

    log(f"Found {len(csv_files)} CSV file(s), parsing...")

    all_rows = []
    for cf in csv_files:
        with open(cf, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                all_rows.append(row)

    log(f"Total rows: {len(all_rows):,}")

    # Identify columns
    sample = all_rows[0] if all_rows else {}
    date_col = next((k for k in sample if k.upper() == "DATETIMEINIT"), None)
    if not date_col:
        date_col = next((k for k in sample if "date" in k.lower() and "request" in k.lower()), None)
    if not date_col:
        date_col = next((k for k in sample if "date" in k.lower() and "init" in k.lower()), None)
    if not date_col:
        date_col = next((k for k in sample if "date" in k.lower()), None)
    cat_col = next((k for k in sample if k.upper() == "PROBLEMCODE"), None)
    if not cat_col:
        cat_col = next((k for k in sample if "problem" in k.lower() or "category" in k.lower()), None)
    if not cat_col:
        cat_col = next((k for k in sample if "type" in k.lower()), None)
    status_col = next((k for k in sample if "status" in k.lower()), None)
    hood_col = next((k for k in sample if "neighborhood" in k.lower() or "nhd" in k.lower()), None)
    lat_col = next((k for k in sample if k.lower() in ("latitude", "lat", "y")), None)
    lng_col = next((k for k in sample if k.lower() in ("longitude", "lng", "lon", "long", "x")), None)
    srx_col = next((k for k in sample if k.upper() == "SRX"), None)
    sry_col = next((k for k in sample if k.upper() == "SRY"), None)
    close_date_col = next((k for k in sample if k.upper() == "DATETIMECLOSED"), None)
    if not close_date_col:
        close_date_col = next((k for k in sample if "close" in k.lower() and "date" in k.lower()), None)

    coords_src = "lat/lon" if (lat_col and lng_col) else ("SRX/SRY" if (srx_col and sry_col) else "none")
    log(f"Columns: date={date_col}, category={cat_col}, status={status_col}, hood={hood_col}, coords={coords_src}")

    def parse_date(s: str) -> str | None:
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    def parse_datetime(s: str) -> datetime | None:
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s.strip(), fmt)
            except ValueError:
                continue
        return None

    # Filter to target year
    year_rows = []
    for row in all_rows:
        d = parse_date(row.get(date_col, "")) if date_col else None
        if d and d.startswith(str(YEAR)):
            year_rows.append(row)

    log(f"Rows for {YEAR}: {len(year_rows):,}")

    if not year_rows:
        log(f"WARNING: No rows found for year {YEAR}. Using all data instead.")
        year_rows = all_rows

    # Aggregations (year-filtered for analytics)
    categories = Counter()
    daily_counts = Counter()
    hourly = Counter()
    weekday = Counter()
    monthly = defaultdict(Counter)
    neighborhoods = defaultdict(lambda: {
        "name": "", "total": 0, "closed": 0, "avgResolutionDays": 0,
        "topCategories": Counter(), "_resolution_days": [],
    })

    for row in year_rows:
        cat = (row.get(cat_col, "") if cat_col else "").strip() or "Unknown"
        categories[cat] += 1

        date_str = parse_date(row.get(date_col, "")) if date_col else None
        if date_str:
            daily_counts[date_str] += 1
            month_key = date_str[:7]
            monthly[month_key][cat] += 1

        dt = parse_datetime(row.get(date_col, "")) if date_col else None
        if dt:
            hourly[str(dt.hour)] += 1
            weekday[str(dt.weekday())] += 1

        # Neighborhood
        hood_name = (row.get(hood_col, "") if hood_col else "").strip()
        if hood_name:
            nb = neighborhoods[hood_name]
            nb["name"] = hood_name
            nb["total"] += 1
            nb["topCategories"][cat] += 1

            status = (row.get(status_col, "") if status_col else "").strip().lower()
            if "closed" in status or "complete" in status:
                nb["closed"] += 1

            if close_date_col and date_col:
                open_dt = parse_datetime(row.get(date_col, ""))
                close_dt = parse_datetime(row.get(close_date_col, ""))
                if open_dt and close_dt and close_dt > open_dt:
                    days = (close_dt - open_dt).days
                    if days < 365:
                        nb["_resolution_days"].append(days)

    # Heatmap points — ALL years for time slider scrubbing
    heatmap_points = []
    for row in all_rows:
        cat = (row.get(cat_col, "") if cat_col else "").strip() or "Unknown"
        date_str = parse_date(row.get(date_col, "")) if date_col else None
        hood_name = (row.get(hood_col, "") if hood_col else "").strip()

        lat, lng = None, None
        if lat_col and lng_col:
            try:
                lat = float(row.get(lat_col, ""))
                lng = float(row.get(lng_col, ""))
            except (ValueError, TypeError):
                pass
        if lat is None and srx_col and sry_col:
            try:
                sx = float(row.get(srx_col, ""))
                sy = float(row.get(sry_col, ""))
                if sx != 0 and sy != 0:
                    lng, lat = web_mercator_to_lnglat(sx, sy)
            except (ValueError, TypeError):
                pass
        if lat is not None and lng is not None:
            if 38.0 < lat < 39.0 and -91.0 < lng < -89.0:
                heatmap_points.append([lat, lng, cat, date_str or "", hood_name])
    log(f"Heatmap points (all years): {len(heatmap_points):,}")

    # Finalize neighborhoods — key by zero-padded NHD_NUM
    final_hoods = {}
    for key, nb in neighborhoods.items():
        try:
            hood_id = str(int(key)).zfill(2)
        except (ValueError, TypeError):
            hood_id = key  # fallback for non-numeric
        res_days = nb.pop("_resolution_days", [])
        avg_res = round(sum(res_days) / len(res_days), 1) if res_days else 0
        top_cats = dict(nb["topCategories"].most_common(5))
        final_hoods[hood_id] = {
            "name": nb["name"],
            "total": nb["total"],
            "closed": nb["closed"],
            "avgResolutionDays": avg_res,
            "topCategories": top_cats,
        }

    monthly_out = {}
    for month_key, cats in sorted(monthly.items()):
        monthly_out[month_key] = dict(cats.most_common(10))

    csb_data = {
        "year": YEAR,
        "totalRequests": sum(categories.values()),
        "categories": dict(categories.most_common()),
        "neighborhoods": final_hoods,
        "dailyCounts": dict(sorted(daily_counts.items())),
        "hourly": dict(sorted(hourly.items(), key=lambda x: int(x[0]))),
        "weekday": dict(sorted(weekday.items(), key=lambda x: int(x[0]))),
        "heatmapPoints": heatmap_points[::max(1, len(heatmap_points) // 50000)][:50000],
        "monthly": monthly_out,
    }

    out_path = OUT_DIR / f"csb_{YEAR}.json"
    with open(out_path, "w") as f:
        json.dump(csb_data, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({out_path.stat().st_size // 1024}KB)")

    latest_path = OUT_DIR / "csb_latest.json"
    shutil.copy2(out_path, latest_path)
    log(f"Copied to {latest_path.name}")

    # ── trends.json (multi-year) ──
    yearly_monthly = defaultdict(dict)
    yearly_categories = defaultdict(Counter)

    for row in all_rows:
        d = parse_date(row.get(date_col, "")) if date_col else None
        if not d:
            continue
        year = d[:4]
        if year not in (str(YEAR), str(YEAR - 1), str(YEAR - 2)):
            continue
        month_key = d[:7]
        cat = (row.get(cat_col, "") if cat_col else "").strip() or "Unknown"
        yearly_monthly[year][month_key] = yearly_monthly[year].get(month_key, 0) + 1
        yearly_categories[year][cat] += 1

    weather_data = fetch_weather(YEAR)

    trends = {
        "yearlyMonthly": {y: dict(sorted(m.items())) for y, m in sorted(yearly_monthly.items())},
        "yearlyCategories": {y: dict(c.most_common()) for y, c in sorted(yearly_categories.items())},
        "weather": weather_data,
    }

    out_path = OUT_DIR / "trends.json"
    with open(out_path, "w") as f:
        json.dump(trends, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({out_path.stat().st_size // 1024}KB)")


def fetch_weather(year: int) -> dict:
    """Fetch daily weather from Open-Meteo historical API for St. Louis."""
    if requests is None:
        log("WARNING: requests not available, skipping weather")
        return {}
    log("Fetching weather data from Open-Meteo...")
    url = (
        f"https://archive-api.open-meteo.com/v1/archive?"
        f"latitude=38.627&longitude=-90.199"
        f"&start_date={year}-01-01&end_date={year}-12-31"
        f"&daily=temperature_2m_max,temperature_2m_min,precipitation_sum"
        f"&temperature_unit=fahrenheit&precipitation_unit=inch"
        f"&timezone=America/Chicago"
    )
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        daily = data.get("daily", {})
        dates = daily.get("time", [])
        highs = daily.get("temperature_2m_max", [])
        lows = daily.get("temperature_2m_min", [])
        precip = daily.get("precipitation_sum", [])

        weather = {}
        for i, d in enumerate(dates):
            weather[d] = {
                "high": round(highs[i], 1) if highs[i] is not None else None,
                "low": round(lows[i], 1) if lows[i] is not None else None,
                "precip": round(precip[i], 2) if precip[i] is not None else 0,
            }
        log(f"Got {len(weather)} days of weather data")
        return weather
    except Exception as e:
        log(f"WARNING: Could not fetch weather data: {e}")
        return {}


# ── 2. Neighborhood Boundaries ───────────────────────────────────────────────

def process_neighborhoods() -> None:
    """Convert neighborhood shapefiles to GeoJSON (reprojected to WGS84)."""
    import geopandas as gpd

    nhd_dir = RAW_DIR / "neighborhoods"
    require_raw(nhd_dir, "neighborhoods")

    shp_files = list(nhd_dir.rglob("*.shp"))
    if not shp_files:
        sys.exit("No .shp files found in raw/neighborhoods/")

    log(f"Converting {shp_files[0].name} to GeoJSON...")
    gdf = gpd.read_file(shp_files[0])

    if gdf.crs and gdf.crs != "EPSG:4326":
        log(f"Reprojecting from {gdf.crs} to EPSG:4326...")
        gdf = gdf.to_crs(epsg=4326)

    geojson = json.loads(gdf.to_json())
    log(f"{len(geojson['features'])} neighborhood features")

    out_path = OUT_DIR / "neighborhoods.geojson"
    with open(out_path, "w") as f:
        json.dump(geojson, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({out_path.stat().st_size // 1024}KB)")


# ── 3. Transit (GTFS) ───────────────────────────────────────────────────────

def process_gtfs() -> None:
    """Process GTFS feed into stops, routes, shapes, stop_stats."""
    gtfs_dir = RAW_DIR / "gtfs"
    require_raw(gtfs_dir, "GTFS")

    # Also check for the zip file directly
    gtfs_zip = RAW_DIR / "google_transit.zip"
    if gtfs_zip.exists():
        zf = zipfile.ZipFile(gtfs_zip)
    else:
        # Look for txt files directly in the extracted dir
        zf = None

    def open_gtfs_file(name: str):
        if zf and name in zf.namelist():
            return io.TextIOWrapper(zf.open(name), encoding="utf-8-sig")
        path = gtfs_dir / name
        if path.exists():
            return open(path, "r", encoding="utf-8-sig")
        return None

    def has_gtfs_file(name: str) -> bool:
        if zf and name in zf.namelist():
            return True
        return (gtfs_dir / name).exists()

    log("Parsing GTFS feed...")

    # ── stops.geojson ──
    f = open_gtfs_file("stops.txt")
    if not f:
        sys.exit("No stops.txt found in GTFS data")
    reader = csv.DictReader(f)
    stops = list(reader)
    f.close()

    features = []
    for s in stops:
        try:
            lat = float(s.get("stop_lat", 0))
            lon = float(s.get("stop_lon", 0))
        except ValueError:
            continue
        if lat == 0 or lon == 0:
            continue
        features.append({
            "type": "Feature",
            "properties": {
                "stop_id": s.get("stop_id", ""),
                "stop_name": s.get("stop_name", ""),
                "stop_code": s.get("stop_code", ""),
            },
            "geometry": {"type": "Point", "coordinates": [lon, lat]},
        })

    stops_geo = {"type": "FeatureCollection", "features": features}
    out_path = OUT_DIR / "stops.geojson"
    with open(out_path, "w") as f:
        json.dump(stops_geo, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({len(features)} stops, {out_path.stat().st_size // 1024}KB)")

    # ── routes.json ──
    f = open_gtfs_file("routes.txt")
    if f:
        reader = csv.DictReader(f)
        routes_list = []
        for r in reader:
            routes_list.append({
                "route_id": r.get("route_id", ""),
                "route_short_name": r.get("route_short_name", ""),
                "route_long_name": r.get("route_long_name", ""),
                "route_type": int(r.get("route_type", 3)),
                "route_color": f"#{r['route_color']}" if r.get("route_color") else "",
            })
        f.close()

        out_path = OUT_DIR / "routes.json"
        with open(out_path, "w") as f:
            json.dump(routes_list, f, separators=(",", ":"))
        log(f"Wrote {out_path.name} ({len(routes_list)} routes)")

    # ── shapes.geojson ──
    if has_gtfs_file("shapes.txt"):
        f = open_gtfs_file("shapes.txt")
        reader = csv.DictReader(f)
        shape_points = defaultdict(list)
        for row in reader:
            sid = row.get("shape_id", "")
            try:
                lat = float(row.get("shape_pt_lat", 0))
                lon = float(row.get("shape_pt_lon", 0))
                seq = int(row.get("shape_pt_sequence", 0))
            except ValueError:
                continue
            shape_points[sid].append((seq, [lon, lat]))
        f.close()

        shape_to_route = {}
        if has_gtfs_file("trips.txt"):
            f = open_gtfs_file("trips.txt")
            reader = csv.DictReader(f)
            for row in reader:
                sid = row.get("shape_id", "")
                rid = row.get("route_id", "")
                if sid and rid:
                    shape_to_route[sid] = rid
            f.close()

        shape_features = []
        for sid, pts in shape_points.items():
            pts.sort(key=lambda x: x[0])
            coords = [p[1] for p in pts]
            if len(coords) < 2:
                continue
            shape_features.append({
                "type": "Feature",
                "properties": {"shape_id": sid, "route_id": shape_to_route.get(sid, "")},
                "geometry": {"type": "LineString", "coordinates": coords},
            })

        shapes_geo = {"type": "FeatureCollection", "features": shape_features}
        out_path = OUT_DIR / "shapes.geojson"
        with open(out_path, "w") as f:
            json.dump(shapes_geo, f, separators=(",", ":"))
        log(f"Wrote {out_path.name} ({len(shape_features)} shapes, {out_path.stat().st_size // 1024}KB)")

    # ── stop_stats.json ──
    if has_gtfs_file("stop_times.txt"):
        trip_to_route = {}
        if has_gtfs_file("trips.txt"):
            f = open_gtfs_file("trips.txt")
            reader = csv.DictReader(f)
            for row in reader:
                trip_to_route[row.get("trip_id", "")] = row.get("route_id", "")
            f.close()

        stop_trips = defaultdict(set)
        stop_routes = defaultdict(set)

        f = open_gtfs_file("stop_times.txt")
        reader = csv.DictReader(f)
        for row in reader:
            sid = row.get("stop_id", "")
            tid = row.get("trip_id", "")
            stop_trips[sid].add(tid)
            rid = trip_to_route.get(tid, "")
            if rid:
                stop_routes[sid].add(rid)
        f.close()

        stats = {}
        for sid in stop_trips:
            stats[sid] = {
                "trip_count": len(stop_trips[sid]),
                "routes": sorted(stop_routes.get(sid, set())),
            }

        out_path = OUT_DIR / "stop_stats.json"
        with open(out_path, "w") as f:
            json.dump(stats, f, separators=(",", ":"))
        log(f"Wrote {out_path.name} ({len(stats)} stops, {out_path.stat().st_size // 1024}KB)")

    if zf:
        zf.close()


# ── 4. Food Desert Tracts ────────────────────────────────────────────────────

def process_food_deserts() -> None:
    """Merge USDA food access data with Census TIGER tract geometries."""
    xlsx_path = RAW_DIR / "food-access-research-atlas-data-download-2019.xlsx"
    require_raw(xlsx_path, "USDA food access")

    tiger_dir = RAW_DIR / "tiger_tracts"
    require_raw(tiger_dir, "TIGER tracts")

    log("Parsing USDA Food Access Research Atlas...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Food Access Research Atlas"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    tract_col = next((i for i, h in enumerate(headers) if h and "CensusTract" in str(h)), None)
    pop_col = next((i for i, h in enumerate(headers) if h and str(h).strip() == "POP2010"), None)
    poverty_col = next((i for i, h in enumerate(headers) if h and "Poverty" in str(h) and "Rate" in str(h)), None)
    lila_col = next((i for i, h in enumerate(headers) if h and str(h).strip() in ("LILATracts_1And10", "LILA1and10")), None)
    vehicle_col = next((i for i, h in enumerate(headers) if h and "lahunv" in str(h).lower()), None)
    income_col = next((i for i, h in enumerate(headers) if h and "Median" in str(h) and "Income" in str(h)), None)

    stl_tracts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tract_id = str(row[tract_col]) if tract_col is not None and row[tract_col] else ""
        if not tract_id.startswith(STL_COUNTY_FIPS):
            continue

        pop = safe_float(row[pop_col] if pop_col is not None else 0)
        poverty = safe_float(row[poverty_col] if poverty_col is not None else 0)
        lila = bool(row[lila_col]) if lila_col is not None and row[lila_col] is not None else False
        vehicle_pct = 0
        if vehicle_col is not None and row[vehicle_col] and pop > 0:
            vehicle_pct = round(safe_float(row[vehicle_col]) / pop * 100, 1)
        median_income = safe_int(row[income_col] if income_col is not None else 0)

        stl_tracts[tract_id] = {
            "pop": safe_int(pop),
            "poverty_rate": round(poverty, 1),
            "lila": lila,
            "pct_no_vehicle": vehicle_pct,
            "median_income": median_income,
        }

    wb.close()
    log(f"Found {len(stl_tracts)} St. Louis census tracts in USDA data")

    shp_files = list(tiger_dir.rglob("*.shp"))
    if not shp_files:
        sys.exit("No .shp files in raw/tiger_tracts/")

    log("Reading TIGER/Line tract geometries...")
    sf = shapefile.Reader(str(shp_files[0]))
    fields = [f[0] for f in sf.fields[1:]]
    geoid_idx = fields.index("GEOID") if "GEOID" in fields else 0
    name_idx = fields.index("NAMELSAD") if "NAMELSAD" in fields else 1

    features = []
    for sr in sf.shapeRecords():
        geoid = str(sr.record[geoid_idx])
        if not geoid.startswith(STL_COUNTY_FIPS):
            continue

        tract_data = stl_tracts.get(geoid, {})
        name = sr.record[name_idx]

        features.append({
            "type": "Feature",
            "properties": {
                "tract_id": geoid,
                "name": name,
                "poverty_rate": tract_data.get("poverty_rate", 0),
                "pop": tract_data.get("pop", 0),
                "pct_no_vehicle": tract_data.get("pct_no_vehicle", 0),
                "nearest_grocery_miles": 1.5,  # placeholder — app computes at runtime
                "lila": tract_data.get("lila", False),
                "median_income": tract_data.get("median_income", 0),
            },
            "geometry": sr.shape.__geo_interface__,
        })

    food_geo = {"type": "FeatureCollection", "features": features}
    out_path = OUT_DIR / "food_deserts.geojson"
    with open(out_path, "w") as f:
        json.dump(food_geo, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({len(features)} tracts, {out_path.stat().st_size // 1024}KB)")


# ── 5. Grocery Stores (embedded) ─────────────────────────────────────────────

GROCERY_STORES = [
    {"name": "Schnucks Arsenal", "chain": "Schnucks", "coords": [-90.2418, 38.5937]},
    {"name": "Schnucks Loughborough", "chain": "Schnucks", "coords": [-90.2689, 38.5830]},
    {"name": "Schnucks Hampton Village", "chain": "Schnucks", "coords": [-90.2920, 38.5917]},
    {"name": "Schnucks Lindbergh", "chain": "Schnucks", "coords": [-90.3360, 38.5803]},
    {"name": "Schnucks Clayton Rd", "chain": "Schnucks", "coords": [-90.3260, 38.6337]},
    {"name": "Schnucks Kirkwood", "chain": "Schnucks", "coords": [-90.4150, 38.5790]},
    {"name": "Schnucks Des Peres", "chain": "Schnucks", "coords": [-90.3990, 38.5970]},
    {"name": "Schnucks Shrewsbury", "chain": "Schnucks", "coords": [-90.3190, 38.5880]},
    {"name": "Dierbergs Brentwood", "chain": "Dierbergs", "coords": [-90.3490, 38.6193]},
    {"name": "Dierbergs Maplewood", "chain": "Dierbergs", "coords": [-90.3160, 38.6100]},
    {"name": "Aldi Gravois", "chain": "Aldi", "coords": [-90.2630, 38.5850]},
    {"name": "Aldi Hampton", "chain": "Aldi", "coords": [-90.2900, 38.6030]},
    {"name": "Aldi Grand", "chain": "Aldi", "coords": [-90.2180, 38.6290]},
    {"name": "Aldi Natural Bridge", "chain": "Aldi", "coords": [-90.2530, 38.6700]},
    {"name": "Save-A-Lot Gravois", "chain": "Save-A-Lot", "coords": [-90.2470, 38.5960]},
    {"name": "Save-A-Lot North Broadway", "chain": "Save-A-Lot", "coords": [-90.1990, 38.6750]},
    {"name": "Save-A-Lot Chippewa", "chain": "Save-A-Lot", "coords": [-90.2730, 38.5900]},
    {"name": "Ruler Foods Lemay Ferry", "chain": "Ruler Foods", "coords": [-90.2800, 38.5430]},
    {"name": "Fresh Thyme Brentwood", "chain": "Fresh Thyme", "coords": [-90.3480, 38.6170]},
    {"name": "Trader Joe's Brentwood", "chain": "Trader Joe's", "coords": [-90.3510, 38.6180]},
    {"name": "Whole Foods Brentwood", "chain": "Whole Foods", "coords": [-90.3520, 38.6160]},
    {"name": "Tower Grove Farmers Market", "chain": "Farmers Market", "coords": [-90.2520, 38.6100]},
    {"name": "Soulard Farmers Market", "chain": "Farmers Market", "coords": [-90.2080, 38.6130]},
]


def write_grocery_stores() -> None:
    """Write embedded grocery store data as GeoJSON."""
    features = [
        {
            "type": "Feature",
            "properties": {"name": s["name"], "chain": s["chain"]},
            "geometry": {"type": "Point", "coordinates": s["coords"]},
        }
        for s in GROCERY_STORES
    ]
    geo = {"type": "FeatureCollection", "features": features}
    out_path = OUT_DIR / "grocery_stores.geojson"
    with open(out_path, "w") as f:
        json.dump(geo, f, indent=2)
    log(f"Wrote {out_path.name} ({len(features)} stores)")


# ── 6. Crime Data (SLMPD) ──────────────────────────────────────────────────

def process_crime() -> None:
    """Process SLMPD crime CSVs from raw data."""
    crime_dir = RAW_DIR / "crime"
    if not crime_dir.exists():
        log("No crime data found in raw/crime/ — skipping")
        return

    csv_files = list(crime_dir.rglob("*.csv")) + list(crime_dir.rglob("*.CSV"))
    if not csv_files:
        log("No CSV files found in raw/crime/ — skipping")
        return

    log(f"Found {len(csv_files)} crime CSV file(s), parsing...")

    all_rows = []
    for cf in csv_files:
        with open(cf, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                all_rows.append(row)

    log(f"Total crime rows: {len(all_rows):,}")

    # Identify columns (SLMPD NIBRS format)
    sample = all_rows[0] if all_rows else {}
    date_col = next((k for k in sample if k.upper() in ("DATEOCCUR", "DATE_OCCUR", "DATEOCCURRED")), None)
    if not date_col:
        date_col = next((k for k in sample if "date" in k.lower()), None)
    crime_col = next((k for k in sample if k.upper() in ("CRIME", "OFFENSE", "NIBRS")), None)
    if not crime_col:
        crime_col = next((k for k in sample if "crime" in k.lower() or "offense" in k.lower()), None)
    desc_col = next((k for k in sample if k.upper() == "DESCRIPTION"), None)
    if not desc_col:
        desc_col = next((k for k in sample if "desc" in k.lower()), None)
    hood_col = next((k for k in sample if k.upper() in ("NEIGHBORHOOD", "NBRHD")), None)
    if not hood_col:
        hood_col = next((k for k in sample if "neighborhood" in k.lower()), None)
    hood_num_col = next((k for k in sample if k.upper() in ("NBHDNUM", "NEIGHBORHOODNUM", "NHD_NUM")), None)
    if not hood_num_col:
        hood_num_col = next((k for k in sample if "nbhd" in k.lower() and "num" in k.lower()), None)
    lat_col = next((k for k in sample if k.upper() in ("XLAT", "LAT", "LATITUDE")), None)
    lng_col = next((k for k in sample if k.upper() in ("XLON", "LON", "LONGITUDE", "LONG")), None)
    fel_col = next((k for k in sample if k.upper() in ("FELMISCIT", "CRIME_TYPE")), None)
    firearm_col = next((k for k in sample if k.upper() in ("FIREARMUSED", "FIREARM")), None)
    district_col = next((k for k in sample if k.upper() == "DISTRICT"), None)

    log(f"Columns: date={date_col}, crime={crime_col}, hood={hood_col}, hoodNum={hood_num_col}, lat={lat_col}")

    def parse_date(s: str) -> str | None:
        if not s:
            return None
        for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    def parse_datetime(s: str) -> datetime | None:
        if not s:
            return None
        for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s.strip(), fmt)
            except ValueError:
                continue
        return None

    # Filter to target year
    year_rows = []
    for row in all_rows:
        d = parse_date(row.get(date_col, "")) if date_col else None
        if d and d.startswith(str(YEAR)):
            year_rows.append(row)

    log(f"Crime rows for {YEAR}: {len(year_rows):,}")

    if not year_rows:
        log(f"WARNING: No crime rows for year {YEAR}. Using all data.")
        year_rows = all_rows

    # Aggregations
    categories = Counter()
    daily_counts = Counter()
    hourly = Counter()
    weekday_counts = Counter()
    monthly = defaultdict(Counter)
    neighborhoods = defaultdict(lambda: {
        "name": "", "total": 0, "topOffenses": Counter(),
        "felonies": 0, "firearmIncidents": 0,
    })
    heatmap_points = []
    total_felonies = 0
    total_firearms = 0

    for row in year_rows:
        # Use description if available, else crime code
        offense = ""
        if desc_col:
            offense = (row.get(desc_col, "") or "").strip()
        if not offense and crime_col:
            offense = (row.get(crime_col, "") or "").strip()
        offense = offense or "Unknown"
        categories[offense] += 1

        date_str = parse_date(row.get(date_col, "")) if date_col else None
        if date_str:
            daily_counts[date_str] += 1
            month_key = date_str[:7]
            monthly[month_key][offense] += 1

        dt = parse_datetime(row.get(date_col, "")) if date_col else None
        if dt:
            hourly[str(dt.hour)] += 1
            weekday_counts[str(dt.weekday())] += 1

        # Felony / firearm tracking
        is_felony = False
        if fel_col:
            fel_val = (row.get(fel_col, "") or "").strip().upper()
            is_felony = fel_val.startswith("FEL")
            if is_felony:
                total_felonies += 1

        has_firearm = False
        if firearm_col:
            fa_val = (row.get(firearm_col, "") or "").strip().upper()
            has_firearm = fa_val in ("Y", "YES", "TRUE", "1")
            if has_firearm:
                total_firearms += 1

        # Neighborhood
        hood_name = (row.get(hood_col, "") if hood_col else "").strip()
        hood_num = (row.get(hood_num_col, "") if hood_num_col else "").strip()
        hood_key = hood_num if hood_num else hood_name
        if hood_key:
            nb = neighborhoods[hood_key]
            nb["name"] = hood_name or hood_key
            nb["total"] += 1
            nb["topOffenses"][offense] += 1
            if is_felony:
                nb["felonies"] += 1
            if has_firearm:
                nb["firearmIncidents"] += 1

        # Heatmap point
        lat, lng = None, None
        if lat_col and lng_col:
            try:
                lat = float(row.get(lat_col, ""))
                lng = float(row.get(lng_col, ""))
            except (ValueError, TypeError):
                pass
        if lat is not None and lng is not None:
            if 38.0 < lat < 39.0 and -91.0 < lng < -89.0:
                hood_id_for_heatmap = str(int(hood_num)).zfill(2) if hood_num and hood_num.isdigit() else hood_name
                heatmap_points.append([lat, lng, offense, date_str or "", hood_id_for_heatmap])

    # Finalize neighborhoods — key by zero-padded NHD_NUM
    final_hoods = {}
    for key, nb in neighborhoods.items():
        # Try to use the hood_num as key, zero-padded
        try:
            nhd_id = str(int(key)).zfill(2)
        except (ValueError, TypeError):
            # If key is a name, use a hash-based assignment
            nhd_id = key
        final_hoods[nhd_id] = {
            "name": nb["name"],
            "total": nb["total"],
            "topOffenses": dict(nb["topOffenses"].most_common(5)),
            "felonies": nb["felonies"],
            "firearmIncidents": nb["firearmIncidents"],
        }

    monthly_out = {}
    for month_key, cats in sorted(monthly.items()):
        monthly_out[month_key] = dict(cats.most_common(10))

    crime_data = {
        "year": YEAR,
        "totalIncidents": sum(categories.values()),
        "totalFelonies": total_felonies,
        "totalFirearms": total_firearms,
        "categories": dict(categories.most_common()),
        "neighborhoods": final_hoods,
        "dailyCounts": dict(sorted(daily_counts.items())),
        "hourly": dict(sorted(hourly.items(), key=lambda x: int(x[0]))),
        "weekday": dict(sorted(weekday_counts.items(), key=lambda x: int(x[0]))),
        "monthly": monthly_out,
        "heatmapPoints": heatmap_points[:50000],
    }

    out_path = OUT_DIR / "crime.json"
    with open(out_path, "w") as f:
        json.dump(crime_data, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({out_path.stat().st_size // 1024}KB)")


# ── 7. ARPA Fund Expenditures ──────────────────────────────────────────────

def process_arpa() -> None:
    """Process ARPA expenditures JSON into frontend-ready format."""
    arpa_path = RAW_DIR / "arpa.json"
    if not arpa_path.exists():
        log("No ARPA data found — skipping")
        return

    with open(arpa_path, "r") as f:
        raw = json.load(f)

    # Handle both array and object responses
    records = raw if isinstance(raw, list) else raw.get("data", raw.get("DATA", []))
    if not isinstance(records, list):
        log("Unexpected ARPA data format — skipping")
        return

    log(f"Processing {len(records)} ARPA transactions...")

    # Category keywords for classification
    CATEGORY_KEYWORDS = {
        "Health": ["health", "covid", "vaccine", "medical", "hospital", "clinic"],
        "Public Safety": ["police", "fire", "safety", "enforcement", "security"],
        "Infrastructure": ["infrastructure", "water", "sewer", "road", "bridge", "building"],
        "Housing": ["housing", "rent", "mortgage", "homeless", "shelter"],
        "Economic Recovery": ["business", "economic", "workforce", "employment", "job"],
        "Community": ["community", "youth", "education", "park", "recreation"],
        "Technology": ["technology", "broadband", "internet", "digital"],
    }

    def categorize(title: str) -> str:
        title_lower = title.lower()
        for cat, keywords in CATEGORY_KEYWORDS.items():
            if any(kw in title_lower for kw in keywords):
                return cat
        return "Other"

    # Parse transactions
    project_totals = defaultdict(lambda: {"title": "", "totalSpent": 0, "category": ""})
    vendor_totals = Counter()
    monthly_spending = Counter()
    total_spent = 0

    for rec in records:
        # Try various field name formats
        amount = safe_float(rec.get("AMOUNT", rec.get("amount", 0)))
        title = str(rec.get("PROJECTTITLE", rec.get("projecttitle", rec.get("PROJECT_TITLE", "")))).strip()
        project_id = str(rec.get("PROJECTID", rec.get("projectid", rec.get("PROJECT_ID", 0))))
        vendor = str(rec.get("VENDOR", rec.get("vendor", ""))).strip()
        date_str = str(rec.get("DATE", rec.get("date", rec.get("EXPENDITURE_DATE", "")))).strip()

        total_spent += amount

        # Project aggregation
        p = project_totals[project_id]
        p["title"] = title or p["title"]
        p["totalSpent"] += amount
        p["category"] = categorize(p["title"])

        # Vendor
        if vendor:
            vendor_totals[vendor] += amount

        # Monthly
        if date_str:
            for fmt in ("%B, %d %Y %H:%M:%S", "%B, %d %Y", "%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    dt = datetime.strptime(date_str, fmt)
                    month_key = dt.strftime("%Y-%m")
                    monthly_spending[month_key] += amount
                    break
                except ValueError:
                    continue

    # Build projects list
    projects = []
    for pid, p in project_totals.items():
        try:
            pid_int = int(pid)
        except (ValueError, TypeError):
            pid_int = 0
        projects.append({
            "id": pid_int,
            "title": p["title"],
            "totalSpent": round(p["totalSpent"], 2),
            "category": p["category"],
        })
    projects.sort(key=lambda x: -x["totalSpent"])

    # Top vendors
    top_vendors = [
        {"name": name, "totalSpent": round(amount, 2)}
        for name, amount in vendor_totals.most_common(20)
    ]

    # Category breakdown
    cat_breakdown = Counter()
    for p in projects:
        cat_breakdown[p["category"]] += p["totalSpent"]
    category_breakdown = {k: round(v, 2) for k, v in cat_breakdown.most_common()}

    # Monthly + cumulative spending
    monthly_sorted = dict(sorted(monthly_spending.items()))
    cumulative = {}
    running = 0
    for k, v in monthly_sorted.items():
        running += v
        cumulative[k] = round(running, 2)
    monthly_sorted = {k: round(v, 2) for k, v in monthly_sorted.items()}

    arpa_data = {
        "totalSpent": round(total_spent, 2),
        "transactionCount": len(records),
        "projects": projects[:100],  # Top 100 projects
        "monthlySpending": monthly_sorted,
        "cumulativeSpending": cumulative,
        "topVendors": top_vendors,
        "categoryBreakdown": category_breakdown,
    }

    out_path = OUT_DIR / "arpa.json"
    with open(out_path, "w") as f:
        json.dump(arpa_data, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({out_path.stat().st_size // 1024}KB)")


# ── 8. Census Demographics ──────────────────────────────────────────────────

def process_demographics() -> None:
    """Process scraped neighborhood census data into demographics JSON.

    The scraped text has a line-by-line table format where the value
    appears on the line AFTER the label:

        Total Population
        7,734
        White alone
        3,334
    """
    demo_path = RAW_DIR / "demographics.json"
    if not demo_path.exists():
        log("No demographics data found — skipping")
        return

    with open(demo_path, "r") as f:
        raw = json.load(f)

    log(f"Processing demographics for {len(raw)} neighborhoods...")

    def parse_int(s: str) -> int:
        """Parse a comma-formatted integer string."""
        try:
            return int(s.strip().replace(",", ""))
        except (ValueError, TypeError):
            return 0

    def build_line_map(text: str) -> dict[str, int]:
        """Build label→value map from consecutive line pairs.

        Scans for lines where the next non-empty line is a number.
        """
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        result: dict[str, int] = {}
        for i in range(len(lines) - 1):
            label = lines[i]
            next_line = lines[i + 1]
            # Next line must look like a number (digits, commas, optional negative)
            if re.match(r"^-?[\d,]+$", next_line):
                result[label.lower()] = parse_int(next_line)
        return result

    demographics = {}
    for nhd_id, page_data in raw.items():
        name = page_data.get("name", f"Neighborhood {nhd_id}")
        # Strip " Census Data" suffix from name
        name = re.sub(r"\s*Census Data\s*$", "", name, flags=re.IGNORECASE).strip()
        text = page_data.get("text", "")

        lm = build_line_map(text)

        # Population (2020)
        pop_2020 = lm.get("total population", 0)

        # Population (2010) — from separate scrape if available
        text_2010 = page_data.get("text_2010", "")
        lm_2010 = build_line_map(text_2010) if text_2010 else {}
        pop_2010 = lm_2010.get("total population", 0)

        # Population change 2010 → 2020
        pop_change = 0.0
        if pop_2010 > 0:
            pop_change = round((pop_2020 - pop_2010) / pop_2010 * 100, 1)

        # Race (use "alone" variants from the total population section)
        white = lm.get("white alone", 0)
        black = lm.get("black or african-american alone", 0)
        asian = lm.get("asian-american alone", 0) or lm.get("asian alone", 0)
        hispanic = lm.get("hispanic or latino", 0)
        other = lm.get("some other race alone", 0) + lm.get("two or more races", 0)

        # Housing
        total_units = lm.get("total housing units", 0)
        occupied = lm.get("occupied housing units", 0)
        vacant = lm.get("vacant housing units", 0)

        vacancy_rate = round(vacant / total_units * 100, 1) if total_units > 0 else 0

        demographics[nhd_id] = {
            "name": name,
            "population": {
                "2020": pop_2020,
                "2010": pop_2010,
                "2000": 0,
            },
            "race": {
                "white": white,
                "black": black,
                "asian": asian,
                "hispanic": hispanic,
                "other": other,
            },
            "housing": {
                "totalUnits": total_units,
                "occupied": occupied,
                "vacant": vacant,
                "vacancyRate": vacancy_rate,
                "ownerOccupied": 0,
                "renterOccupied": 0,
            },
            "popChange10to20": pop_change,
        }

    out_path = OUT_DIR / "demographics.json"
    with open(out_path, "w") as f:
        json.dump(demographics, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({len(demographics)} neighborhoods, {out_path.stat().st_size // 1024}KB)")


# ── 9. Real Vacancy Data ───────────────────────────────────────────────────

def process_vacancies() -> None:
    """Join vacancy API overview with parcel shapefile to produce vacancies.json.

    Data sources:
    - raw/vacancies/vacancy_overview.json — from stlcitypermits.com API (keyed by HANDLE)
    - raw/parcels/PARCELS/PARCELS.shp — city parcel shapefile (has geometry, address, owner, etc.)

    The vacancy overview provides: minor violations, major violations, CSB complaints, unpaid fines.
    The parcel shapefile provides: address, owner, lat/lng (via centroid), neighborhood, lot size, etc.
    We join on HANDLE and compute triage scores from real data.
    """
    import geopandas as gpd

    vacancy_dir = RAW_DIR / "vacancies"
    overview_path = vacancy_dir / "vacancy_overview.json"
    if not overview_path.exists():
        log("No vacancy_overview.json found — skipping (run fetch_raw.py first)")
        return

    parcel_dir = RAW_DIR / "parcels"
    shp_files = list(parcel_dir.rglob("*.shp")) if parcel_dir.exists() else []
    if not shp_files:
        log("No parcel shapefile found — skipping (run fetch_raw.py first)")
        return

    # Load vacancy overview (HANDLE -> {mo, vmin, vmaj, csb, unpd})
    with open(overview_path, "r") as f:
        vacancy_overview = json.load(f)
    log(f"Loaded {len(vacancy_overview)} vacant parcels from API")

    # Load parcel shapefile and reproject to WGS84
    log(f"Reading parcel shapefile ({shp_files[0].name})...")
    gdf = gpd.read_file(shp_files[0])
    if gdf.crs and gdf.crs != "EPSG:4326":
        log(f"Reprojecting from {gdf.crs} to EPSG:4326...")
        gdf = gdf.to_crs(epsg=4326)

    # Index parcels by HANDLE for fast lookup
    parcel_lookup = {}
    for _, row in gdf.iterrows():
        handle = str(row.get("HANDLE", "")).strip()
        if handle:
            parcel_lookup[handle] = row
    log(f"Indexed {len(parcel_lookup)} parcels by HANDLE")

    # Triage score weights (mirrors scoring.ts)
    WEIGHTS = {
        "condition": 0.25,
        "complaintDensity": 0.2,
        "lotSize": 0.1,
        "ownership": 0.15,
        "proximity": 0.15,
        "taxDelinquency": 0.15,
    }

    properties = []
    matched = 0
    for handle, vdata in vacancy_overview.items():
        parcel = parcel_lookup.get(handle)
        if parcel is None:
            continue
        matched += 1

        # Extract centroid lat/lng from geometry
        geom = parcel.get("geometry")
        if geom is None or geom.is_empty:
            continue
        centroid = geom.centroid
        lat = round(centroid.y, 6)
        lng = round(centroid.x, 6)
        if not (38.0 < lat < 39.0 and -91.0 < lng < -89.0):
            continue

        # Parcel fields
        address = str(parcel.get("SITEADDR", "")).strip()
        if not address:
            address = f"{parcel.get('LowAddrNum', '')} {parcel.get('StName', '')} {parcel.get('StType', '')}".strip()
        owner_raw = str(parcel.get("OWNERNAME", "")).strip().upper()
        ward = safe_int(parcel.get("WARD", 0))
        nbrhd = safe_int(parcel.get("NBRHD", 0))
        neighborhood = str(nbrhd).zfill(2) if nbrhd > 0 else ""
        zip_code = str(safe_int(parcel.get("ZIP", 0)))
        lot_sqft = safe_int(parcel.get("SQFT", 0)) or 3000
        zoning = str(parcel.get("Zoning", "B")).strip() or "B"
        assessed_value = safe_float(parcel.get("AsdTotal", 0))
        tax_balance = safe_float(parcel.get("TaxBalance", 0))
        year_built = safe_int(parcel.get("FirstYearB", 0)) or None
        num_bldgs = safe_int(parcel.get("NbrOfBldgs", 0))
        vacant_lot = safe_int(parcel.get("VacantLot", 0))
        parcel_id = str(parcel.get("ParcelId", handle)).strip()

        # Determine owner type
        if "LRA" in owner_raw or "LAND REUTILIZATION" in owner_raw:
            owner = "LRA"
        elif "CITY" in owner_raw or "ST. LOUIS" in owner_raw or "SAINT LOUIS" in owner_raw:
            owner = "CITY"
        else:
            owner = "PRIVATE"

        property_type = "lot" if (vacant_lot == 1 or num_bldgs == 0) else "building"

        # Violation data from API
        minor_violations = safe_int(vdata.get("vmin", 0))
        major_violations = safe_int(vdata.get("vmaj", 0))
        csb_complaints = safe_int(vdata.get("csb", 0))
        unpaid_fines = safe_float(vdata.get("unpd", 0))
        total_violations = minor_violations + major_violations

        # Derive condition rating (1=worst, 5=best) from violations
        if major_violations >= 10:
            condition = 1
        elif major_violations >= 5:
            condition = 2
        elif major_violations >= 2 or total_violations >= 8:
            condition = 3
        elif total_violations >= 1:
            condition = 4
        else:
            condition = 5

        # Estimate tax delinquency from balance
        est_annual_tax = max(assessed_value * 0.08, 500) if assessed_value > 0 else 500
        tax_years = min(10, round(tax_balance / est_annual_tax)) if tax_balance > 0 else 0

        # Proximity score: use CSB complaints as a proxy for neighborhood activity
        proximity = min(100, 30 + csb_complaints * 15)
        complaints_nearby = csb_complaints

        # Score breakdown (mirrors scoring.ts)
        scores = {}
        scores["condition"] = round(((5 - condition) / 4) * 100)
        scores["complaintDensity"] = min(100, round((total_violations / 20) * 100))
        scores["lotSize"] = round(min(lot_sqft / 10000, 1) * 100)
        if owner == "LRA":
            scores["ownership"] = 100
        elif owner == "CITY":
            scores["ownership"] = 70
        else:
            scores["ownership"] = min(100, round((tax_years / 5) * 50))
        scores["proximity"] = proximity
        scores["taxDelinquency"] = min(100, round((tax_years / 10) * 100))

        composite = round(
            scores["condition"] * WEIGHTS["condition"]
            + scores["complaintDensity"] * WEIGHTS["complaintDensity"]
            + scores["lotSize"] * WEIGHTS["lotSize"]
            + scores["ownership"] * WEIGHTS["ownership"]
            + scores["proximity"] * WEIGHTS["proximity"]
            + scores["taxDelinquency"] * WEIGHTS["taxDelinquency"]
        )
        triage_score = min(100, max(0, composite))

        # Best use determination
        housing_fit = (35 if property_type == "building" else 0) + condition * 8 + (proximity / 100) * 25
        solar_fit = (30 if property_type == "lot" else 0) + min(40, (lot_sqft / 15000) * 40) + (15 if owner == "LRA" else 0)
        garden_fit = (25 if property_type == "lot" else 0) + (30 if 2000 <= lot_sqft <= 8000 else 15) + (proximity / 100) * 25
        best = max(housing_fit, solar_fit, garden_fit)
        if best == housing_fit:
            best_use = "housing"
        elif best == solar_fit:
            best_use = "solar"
        else:
            best_use = "garden"

        properties.append({
            "id": matched,
            "parcelId": parcel_id,
            "address": address,
            "zip": zip_code,
            "lat": lat,
            "lng": lng,
            "ward": ward,
            "neighborhood": neighborhood,
            "propertyType": property_type,
            "owner": owner,
            "conditionRating": condition,
            "lotSqFt": lot_sqft,
            "zoning": zoning,
            "taxYearsDelinquent": tax_years,
            "complaintsNearby": complaints_nearby,
            "proximityScore": proximity,
            "neighborhoodDemand": 50,
            "boardUpStatus": "Unknown",
            "violationCount": total_violations,
            "condemned": major_violations >= 10,
            "assessedValue": assessed_value,
            "yearBuilt": year_built,
            "stories": 1,
            "recentComplaints": [],
            "vacancyCategory": "Vacant Building",
            "triageScore": triage_score,
            "scoreBreakdown": scores,
            "bestUse": best_use,
        })

    log(f"Matched {matched} of {len(vacancy_overview)} vacant parcels to parcel shapefile")

    out_path = OUT_DIR / "vacancies.json"
    with open(out_path, "w") as f:
        json.dump(properties, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({len(properties)} properties, {out_path.stat().st_size // 1024}KB)")


# ── 10. Census ACS Housing Data ──────────────────────────────────────────────

def process_housing() -> None:
    """Spatial-join ACS tract-level housing data to neighborhoods, output housing.json."""
    import geopandas as gpd

    acs_path = RAW_DIR / "housing_acs.json"
    if not acs_path.exists():
        log("No housing_acs.json found — skipping (set CENSUS_API_KEY and run fetch_raw.py)")
        return

    nhd_path = OUT_DIR / "neighborhoods.geojson"
    if not nhd_path.exists():
        log("neighborhoods.geojson not found — run process_neighborhoods first")
        return

    tiger_dir = RAW_DIR / "tiger_tracts"
    shp_files = list(tiger_dir.rglob("*.shp")) if tiger_dir.exists() else []
    if not shp_files:
        log("No TIGER tract shapefile found — skipping housing")
        return

    # Load ACS data (Census API returns header row + data rows)
    with open(acs_path, "r") as f:
        raw = json.load(f)

    headers = raw[0]
    rows = raw[1:]
    log(f"Loaded {len(rows)} ACS tract records")

    # Build GEOID -> {rent, value} lookup
    # GEOID = state + county + tract
    name_idx = headers.index("NAME")
    rent_idx = headers.index("B25064_001E")
    value_idx = headers.index("B25077_001E")
    state_idx = headers.index("state")
    county_idx = headers.index("county")
    tract_idx = headers.index("tract")

    tract_data = {}
    rents = []
    values = []
    for row in rows:
        geoid = row[state_idx] + row[county_idx] + row[tract_idx]
        rent = safe_int(row[rent_idx]) if row[rent_idx] not in (None, "", "-666666666") else None
        value = safe_int(row[value_idx]) if row[value_idx] not in (None, "", "-666666666") else None
        tract_data[geoid] = {"rent": rent, "value": value}
        if rent and rent > 0:
            rents.append(rent)
        if value and value > 0:
            values.append(value)

    city_median_rent = sorted(rents)[len(rents) // 2] if rents else None
    city_median_value = sorted(values)[len(values) // 2] if values else None
    log(f"City median rent: ${city_median_rent}, home value: ${city_median_value}")

    # Load TIGER tracts as GeoDataFrame
    log("Loading TIGER tracts for spatial join...")
    tracts_gdf = gpd.read_file(shp_files[0])
    tracts_gdf = tracts_gdf[tracts_gdf["GEOID"].str.startswith(STL_COUNTY_FIPS)]
    if tracts_gdf.crs and tracts_gdf.crs != "EPSG:4326":
        tracts_gdf = tracts_gdf.to_crs(epsg=4326)

    # Add ACS values to tracts
    tracts_gdf["rent"] = tracts_gdf["GEOID"].map(lambda g: (tract_data.get(g, {}).get("rent")))
    tracts_gdf["home_value"] = tracts_gdf["GEOID"].map(lambda g: (tract_data.get(g, {}).get("value")))

    # Load neighborhoods
    nhd_gdf = gpd.read_file(nhd_path)
    if nhd_gdf.crs and nhd_gdf.crs != "EPSG:4326":
        nhd_gdf = nhd_gdf.to_crs(epsg=4326)

    # Spatial join: assign each tract to the neighborhood it overlaps most
    joined = gpd.sjoin(tracts_gdf, nhd_gdf, how="left", predicate="intersects")
    log(f"Spatial join: {len(joined)} tract-neighborhood matches")

    # Aggregate per neighborhood
    neighborhoods = {}
    for nhd_num in nhd_gdf["NHD_NUM"].unique():
        nhd_id = str(int(nhd_num)).zfill(2)
        nhd_row = nhd_gdf[nhd_gdf["NHD_NUM"] == nhd_num].iloc[0]
        nhd_name = nhd_row.get("NHD_NAME", f"Neighborhood {nhd_id}")

        matched = joined[joined["NHD_NUM"] == nhd_num]
        tract_count = len(matched)

        rent_vals = [r for r in matched["rent"].dropna() if r > 0]
        value_vals = [v for v in matched["home_value"].dropna() if v > 0]

        avg_rent = round(sum(rent_vals) / len(rent_vals)) if rent_vals else None
        avg_value = round(sum(value_vals) / len(value_vals)) if value_vals else None

        neighborhoods[nhd_id] = {
            "name": nhd_name,
            "medianRent": avg_rent,
            "medianHomeValue": avg_value,
            "tractCount": tract_count,
        }

    housing = {
        "year": ACS_YEAR,
        "cityMedianRent": city_median_rent,
        "cityMedianHomeValue": city_median_value,
        "neighborhoods": neighborhoods,
    }

    out_path = OUT_DIR / "housing.json"
    with open(out_path, "w") as f:
        json.dump(housing, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({len(neighborhoods)} neighborhoods, {out_path.stat().st_size // 1024}KB)")


# ── Main ─────────────────────────────────────────────────────────────────────

STEPS = {
    "neighborhoods": ("Neighborhoods", process_neighborhoods),
    "gtfs": ("GTFS transit", process_gtfs),
    "food": ("Food deserts", process_food_deserts),
    "grocery": ("Grocery stores", write_grocery_stores),
    "csb": ("CSB 311 data", process_csb),
    "crime": ("Crime data", process_crime),
    "arpa": ("ARPA funds", process_arpa),
    "demographics": ("Demographics", process_demographics),
    "vacancies": ("Vacancy data", process_vacancies),
    "housing": ("Housing (ACS)", process_housing),
}


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Process raw data into frontend JSON")
    parser.add_argument(
        "--only",
        type=str,
        help=f"Process only this step. Choices: {', '.join(STEPS.keys())}",
    )
    parser.add_argument("--list", action="store_true", help="List available steps and exit")
    args = parser.parse_args()

    if args.list:
        print("Available steps:")
        for key, (name, _) in STEPS.items():
            print(f"  {key:<20} {name}")
        return

    print("=" * 60)
    print("  STL Urban Analytics — Data Cleaner")
    print(f"  Raw input:  {RAW_DIR}")
    print(f"  Output:     {OUT_DIR}")
    print(f"  Target year: {YEAR}")
    if args.only:
        print(f"  Only: {args.only}")
    print("=" * 60)

    if not RAW_DIR.exists():
        sys.exit(f"\nNo raw data found at {RAW_DIR}\nRun `uv run python scripts/fetch_raw.py` first.")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if args.only:
        if args.only not in STEPS:
            sys.exit(f"Unknown step '{args.only}'. Use --list to see options.")

    for key, (name, fn) in STEPS.items():
        if args.only and args.only != key:
            continue
        try:
            print(f"\n── {name} ──")
            fn()
        except SystemExit:
            raise
        except Exception as e:
            print(f"\n\u274c {name} failed: {e}")

    # Summary
    print("\n" + "=" * 60)
    print("  Done! Files in public/data/:")
    print("=" * 60)
    for f in sorted(OUT_DIR.iterdir()):
        size = f.stat().st_size
        unit = "KB" if size > 1024 else "B"
        val = size // 1024 if size > 1024 else size
        print(f"  {f.name:<30} {val:>6} {unit}")


if __name__ == "__main__":
    main()
